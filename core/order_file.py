from dataclasses import dataclass, InitVar, field
from datetime import date
from operator import attrgetter
from pathlib import Path
from typing import Tuple, List, Dict

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.table import Table, TableColumn


@dataclass()
class Order:
    """
    Cодержит сведения, необходимые для заказа выписки из ЕГРН на зу и сохранения реквизитов заявки
    """
    cn: str = None
    region: str = None
    type: str = "Земельный участок"
    area: int = None
    district: str = None
    house: str = None
    apartment: str = None
    order_num: str = None
    order_code: str = None
    order_date: date = None
    index: int = None


@dataclass()
class OrderFileXlsx:
    """
    Считывает и записывает заказ (Order) из/в файл excel
    """

    workbook_path: Path = Path('orders.xlsx')
    table_name: InitVar[str] = 'orders'
    ws: worksheet = field(init=False)
    tbl: Table = field(init=False)

    def __post_init__(self, table_name):
        self.wb = load_workbook(self.workbook_path, data_only=True)
        self.ws, self.tbl = [(ws, ws.tables[table_name]) for ws in self.wb.worksheets
                             for tbl in ws.tables.keys() if table_name == tbl][0]

    def get_tbl_cols_and_data(self) -> Tuple[List[TableColumn], Tuple[Tuple[Cell]]]:
        cols: List[TableColumn] = self.tbl.tableColumns
        data: Tuple[Tuple[Cell]] = self.ws[self.tbl.ref][1:]
        return cols, data

    def get_order_result_cols_index(self) -> Dict[str, str]:
        cols, data = self.get_tbl_cols_and_data()
        num = [i for i, c in enumerate(cols) if c.name == "order_num"][0]
        code = [i for i, c in enumerate(cols) if c.name == "order_code"][0]
        date_order = [i for i, c in enumerate(cols) if c.name == "order_date"][0]

        return {"order_num": data[0][num].column_letter,
                "order_code": data[0][code].column_letter,
                "order_date": data[0][date_order].column_letter}

    def get_orders(self) -> List[Order]:
        cols, data = self.get_tbl_cols_and_data()

        headers = [col.name for col in cols]
        # Загрузка данных из таблицы Excel, кроме заголовка
        orders: List[Order] = []
        for row in data:
            all_data = dict(zip(headers, [c.value for c in row]))
            # Фильтруем заголовки по данным, которые поддерживает класс Order if col.name in Order.__dict__.keys()
            data = {k: v for k, v in all_data.items() if k in Order.__dict__}

            order = Order(**data, index=row[0].row)
            orders.append(order)

        # Фильтрация исполненных заказов и сортировка по типу объекта, республикам и КН
        # чтобы сократить время на выбор республики и района
        orders_fs = sorted((order for order in orders if order.order_num is None),
                           key=attrgetter("type", "region", "cn"))

        return orders_fs

    def write_order(self, order: Order) -> None:
        cols_index = self.get_order_result_cols_index()
        for col in cols_index.keys():
            self.ws[f"{cols_index[col]}{order.index}"] = order.__getattribute__(col)

        self.wb.save(self.workbook_path)
